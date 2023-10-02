


import openpyxl
import time
import sys

ubiAgrupcomb = 'C:/Users/' #colocar ubicacion del archivo fuente
ubiGTi = 'C:/Users/' #colocar ubicacion del archivo GTi

fileAgrupcomb = openpyxl.load_workbook(ubiAgrupcomb)
fileGTi = openpyxl.load_workbook(ubiGTi)

#la celda de la variante de la estacion siempre es la misma celda en todas las hojas del archivo de agrupcomb

def pasar1aVariante(celdaActualHiloAgrupcomb, celdaActualAgrVariante, celdaActualGTi, celdaActualGTiVariante, numeroDeVariantes, nameHojaAgrupcomb, nameHojaGTi):
    hojaGTi = fileGTi[nameHojaGTi]
    hojaAgrupcomb = fileAgrupcomb[nameHojaAgrupcomb]
    
    columnAgrupcomb = int(celdaActualAgrVariante.column)
    columnGTi = int(celdaActualGTiVariante.column)
    celdaActualGTi = hojaGTi.cell(row= int(celdaActualGTi.row), column= columnGTi)
           
    for i in range(0,numeroDeVariantes):
        borrarCelda = hojaGTi.cell(row=int(celdaActualGTi.row), column= columnGTi, value = "")
        print("el valor de la celda: " + str(celdaActualGTi) + " ahora es: " + str(celdaActualAgrVariante.value))
        nuevoValorCrossGTi = hojaGTi.cell(row=int(celdaActualGTi.row), column= columnGTi, value= celdaActualAgrVariante.value)
        columnAgrupcomb +=1
        columnGTi +=1
        
        celdaActualGTi = hojaGTi.cell(row= int(celdaActualGTi.row), column= columnGTi)
        celdaActualAgrVariante = hojaAgrupcomb.cell(row= int(celdaActualAgrVariante.row), column= columnAgrupcomb)
    return        
               
        



def aplicarCrossMultiple(celdaActualGTi, celdaActualGTiVariante, nameHojaAgrupcomb, nameHojaGTi):
    
    hojaGTi = fileGTi[nameHojaGTi]
    hojaAgrupcomb = fileAgrupcomb[nameHojaAgrupcomb]
    
    rowAgrupcombHilo = 10
    columnAgrupcombVariante= 14
    celdaActualHiloAgrupcomb = hojaAgrupcomb.cell(row=rowAgrupcombHilo, column=1)
    
    rowHiloGTi = int(celdaActualGTi.row)
    
    columnOnlyContador = 14
    celdaActualVarianteAgrupcomb = hojaAgrupcomb.cell(row=9, column= columnOnlyContador)
    numeroDeVariantes = 0
    
    while(True):
        if(celdaActualVarianteAgrupcomb.value != None):
            numeroDeVariantes += 1
        else:
            break
        
        columnOnlyContador += 1
        celdaActualVarianteAgrupcomb = hojaAgrupcomb.cell(row=9, column= columnOnlyContador)
    
    print("el numero de variantes totales es: " + str(numeroDeVariantes))
                       
    
    while(celdaActualGTi.value != None):
        
        if (celdaActualGTi.value == "FREE"):
            print("celda FREE")
            rowHiloGTi += 1
            celdaActualGTi = hojaGTi.cell(row=rowHiloGTi, column= celdaActualGTi.column)        
        
        print("buscando el hilo: " + celdaActualGTi.value + " en la celda: " + str(celdaActualHiloAgrupcomb) + " de Agrupcomb")
        
        
        if (celdaActualGTi.value == celdaActualHiloAgrupcomb.value) :
            celdaActualAgrVariante = hojaAgrupcomb.cell(row= int(celdaActualHiloAgrupcomb.row), column=14)
            
            print("hilo: " + celdaActualGTi.value + " encontrado en la celda: " + str(celdaActualHiloAgrupcomb) + " de agrupcomb")
            pasar1aVariante(celdaActualHiloAgrupcomb, celdaActualAgrVariante, celdaActualGTi, celdaActualGTiVariante, numeroDeVariantes, nameHojaAgrupcomb, nameHojaGTi)
            rowHiloGTi += 1
            rowAgrupcombHilo = 10   
            celdaActualGTi = hojaGTi.cell(row=rowHiloGTi, column= celdaActualGTi.column)
            celdaActualHiloAgrupcomb = hojaAgrupcomb.cell(row= 10, column= 1)
        else:
            rowAgrupcombHilo += 1
            celdaActualHiloAgrupcomb = hojaAgrupcomb.cell(row=rowAgrupcombHilo, column=1)
        
        
        
        if (celdaActualHiloAgrupcomb.value == None):
            print("no se encontro hilo, revise configuracion")
            sys.exit(1) 
    
    print("fin")                
    return
    
    



def validarSerie(serieActual, celdaDelHilo, nameHojaGTi):
    hojaGTi = fileGTi[nameHojaGTi]
    celdaInicial = celdaDelHilo
    rowActual = int(celdaDelHilo.row)
    
    while(True):
        if (celdaDelHilo.value == serieActual):
            print("La serie: " + serieActual + " fue encontrada en la celda: " + str(celdaDelHilo))
            return 0
        else:
            print("La serie no fue encontrada en la celda: " + str(celdaDelHilo) + "---buscando en la siguiente celda...")
            rowActual+=1
            celdaDelHilo = hojaGTi.cell(row= rowActual,column= celdaDelHilo.column)
        
        if (celdaDelHilo.value == None):
            print("no se encontro la celda en la columna, buscando en la columna D...")
            celdaDelHilo = hojaGTi.cell(row= celdaInicial.row, column= int(celdaInicial.column) + 1)
            rowActual = int(celdaInicial.row)
            while(True):
                if(celdaDelHilo.value == serieActual):
                    print("La serie: " + serieActual + " fue encontrada en la celda: " + str(celdaDelHilo))
                    print("El valor sera asignado en la misma fila en la columna correspondiente")
                    SerieAgregada = hojaGTi.cell(row=celdaDelHilo.row, column=celdaInicial.column, value= celdaDelHilo.value)
                    return 0
                else:
                    print("La serie no fue encontrada en la celda: " + str(celdaDelHilo) + "---buscando en la siguiente celda...")
                    rowActual += 1
                    celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)
                
                if(celdaDelHilo.value == None):
                    print("no se encontro la celda en la columna, buscando en la Columna B...")
                    celdaDelHilo = hojaGTi.cell(row= celdaInicial.row, column= int(celdaInicial.column) - 1)
                    rowActual = int(celdaInicial.row)
                    while(True):
                        if(celdaDelHilo.value == serieActual):
                            print("La serie: " + serieActual + " fue encontrada en la celda: " + str(celdaDelHilo))
                            print("El valor sera asignado en la misma fila en la columna correspondiente")
                            SerieAgregada = hojaGTi.cell(row=celdaDelHilo.row, column=celdaInicial.column, value= celdaDelHilo.value)
                            return 0
                        else:
                            print("La serie no fue encontrada en la celda: " + str(celdaDelHilo) + "---buscando en la siguiente celda...")
                            rowActual += 1
                            celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)
                         
                        if(celdaDelHilo.value == None):
                            print("no se encontro ningun hilo, revise el agrupcomb y agregue lo necesario")
                            sys.exit(1)       
                                                            
###################################################

def validarTierras(tierraActual,celdaDelHilo, nameHojaGTi):
    
    hojaGTi = fileGTi[nameHojaGTi]
    celdaInicial = celdaDelHilo
    rowActual = int(celdaDelHilo.row)
    
    while(True):
        if (celdaDelHilo.value == tierraActual):
            print("La tierra: " + tierraActual + " fue encontrada en la celda: " + str(celdaDelHilo))
            return 0
        else:
            print("La tierra no fue encontrada en la celda: " + str(celdaDelHilo) + "---buscando en la siguiente celda...")
            rowActual+=1
            celdaDelHilo = hojaGTi.cell(row= rowActual,column= celdaDelHilo.column)
        
        if (celdaDelHilo.value == None):
            print("no se encontro la celda en la columna, buscando en la columna D...")
            celdaDelHilo = hojaGTi.cell(row= celdaInicial.row, column= int(celdaInicial.column) + 1)
            rowActual = int(celdaInicial.row)
            while(True):
                if(celdaDelHilo.value == tierraActual):
                    print("La tierra: " + tierraActual + " fue encontrada en la celda: " + str(celdaDelHilo))
                    print("El valor sera asignado en la misma fila en la columna correspondiente")
                    SerieAgregada = hojaGTi.cell(row=celdaDelHilo.row, column=celdaInicial.column, value= celdaDelHilo.value)
                    rowActual += 1
                    celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)                      
                    
                    while(celdaDelHilo.value != None):
                        print("buscando si existe otra posicion....")      
                        if(celdaDelHilo.value == tierraActual):
                            print("La tierra: " + tierraActual + "tambien se encontro en la celda: " + str(celdaDelHilo))
                            print("el valor sera asignado en la misma fila en la columna correspondiente")
                            serieAgregada = SerieAgregada = hojaGTi.cell(row=celdaDelHilo.row, column=celdaInicial.column, value= celdaDelHilo.value)
                            rowActual += 1
                            celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)       
                            while(celdaDelHilo.value != None):
                                print("buscando si existe otra posicion...")
                                if(celdaDelHilo.value == tierraActual):
                                    print("La tierra: " + tierraActual + "tambien se encontro en la celda: " + str(celdaDelHilo))
                                    print("el valor sera asignado en la misma fila en la columna correspondiente")
                                    serieAgregada = SerieAgregada = hojaGTi.cell(row=celdaDelHilo.row, column=celdaInicial.column, value= celdaDelHilo.value)
                                    return 0
                                else:
                                    rowActual += 1
                                    celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)  
                            return 0    
                        else:
                            rowActual += 1
                            celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)      
                                    
                                
                                
                                
                        
                    return 0
                else:
                    print("La tierra no fue encontrada en la celda: " + str(celdaDelHilo) + "---buscando en la siguiente celda...")
                    rowActual += 1
                    celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)
                
                if(celdaDelHilo.value == None):
                    print("no se encontro la celda en la columna, buscando en la Columna B...")
                    celdaDelHilo = hojaGTi.cell(row= celdaInicial.row, column= int(celdaInicial.column) - 1)
                    rowActual = int(celdaInicial.row)
                    while(True):
                        if(celdaDelHilo.value == tierraActual):
                            print("La tierra: " + tierraActual + " fue encontrada en la celda: " + str(celdaDelHilo))
                            print("El valor sera asignado en la misma fila en la columna correspondiente")
                            SerieAgregada = hojaGTi.cell(row=celdaDelHilo.row, column=celdaInicial.column, value= celdaDelHilo.value)
                            return 0
                        else:
                            print("La tierra no fue encontrada en la celda: " + str(celdaDelHilo) + "---buscando en la siguiente celda...")
                            rowActual += 1
                            celdaDelHilo = hojaGTi.cell(row= rowActual, column=celdaDelHilo.column)
                         
                        if(celdaDelHilo.value == None):
                            print("no se encontro ningun hilo, revise el agrupcomb y agregue lo necesario")
                            sys.exit(1)       




  


#################################
def buscarPrimeraCeldaDeVariante(nameHojaAgrupcomb, nameHojaGTi):
    hojaGTi = fileGTi[nameHojaGTi]
    
    hojaAgrupcomb = fileAgrupcomb[nameHojaAgrupcomb]
  
    varianteActualAgr = hojaAgrupcomb.cell(row=9, column=14)
    print("Buscando variante de estacion: " + varianteActualAgr.value + " en la hoja GTi")    
    columnGti = 3    
    
    for i in range(1,100):
        celdaActualGTi = hojaGTi.cell(row=2, column=columnGti)
        if varianteActualAgr.value == celdaActualGTi.value:
            print("Se encontro la variante de estacion: " + varianteActualAgr.value + "en la celda: " + str(celdaActualGTi))
                       
            return varianteActualAgr, celdaActualGTi
            break
        else:
            print("No se encontro la variante en la celda: " + str(celdaActualGTi) + " --Buscando en siguiente columna")
            columnGti += 1
    
    print("No se encontro variante especificada")
    print("Revise el archivo agrupcomb y asegurese de que esta bien estructurado, o que al menos la primera variante este agregada al archivo GTi")        
    sys.exit(1)


def validarVarianteEnColumna(celdaActualGTi, nameHojaAgrupcomb, nameHojaGTi):
    hojaGTi = fileGTi[nameHojaGTi]
    hojaAgrupcomb = fileAgrupcomb[nameHojaAgrupcomb]
    columnVarAgrupcomb = 14
    columnVarGTi = int(celdaActualGTi.column)
    
    celdaActualAgr = hojaAgrupcomb.cell(row=9, column=columnVarAgrupcomb)
    
    print("la variante a buscar es: " + str(celdaActualAgr.value) + " que se encuentra en la celda: " + str(celdaActualGTi))    
        

    while (celdaActualAgr.value != None):
        if (celdaActualAgr.value == celdaActualGTi.value):
            print("Variante: " + celdaActualAgr.value + " si existe en GTi en la celda: " + str(celdaActualGTi))
        else:
            print("Variante:" + celdaActualAgr.value + " no encontrada en la celda: " + str(celdaActualGTi) )
            hojaGTi.insert_cols(celdaActualGTi.column)
            celdaNueva = hojaGTi.cell(row=2, column=columnVarGTi, value=celdaActualAgr.value)
            print(celdaActualAgr.value + " ha sido agregada al programa GTi en la columna " + str(celdaActualGTi.column))
    
        columnVarAgrupcomb += 1
        columnVarGTi += 1
        celdaActualAgr = hojaAgrupcomb.cell(row=9, column=columnVarAgrupcomb)
        celdaActualGTi = hojaGTi.cell(row=2,column=columnVarGTi)

    return 0


#el siguiente metodo aplica tanto izquierdo como derecho
def BuscarFilaDeHilo(celdaActualGTiVariante, nameHojaAgrupcomb, nameHojaGTi):
    
    hojaGTi = fileGTi[nameHojaGTi]
    hojaAgrupcomb = fileAgrupcomb[nameHojaAgrupcomb]
    rowActual = int(celdaActualGTiVariante.row)
    
    print("Buscando 1 en la columna " + str(celdaActualGTiVariante.column) + "para encontrar Hilo en Fila")
    while(True):
        if celdaActualGTiVariante.value == 1:
            rowHiloEncontrado = int(celdaActualGTiVariante.row)
            print("se encontro el 1 en la fila: " + str(rowHiloEncontrado) + "--- buscando hilo en esa posicion")
            break
        else:
            print("no se encontro 1 en celda: " + str(celdaActualGTiVariante) + "-- buscando en siguiente fila...")
            rowActual += 1
            celdaActualGTiVariante = hojaGTi.cell(row=rowActual, column= int(celdaActualGTiVariante.column))
        
        if rowActual == 400:
            print("no se encontro ningun valor 1, revise el archivo e intente de nuevo")
            sys.exit(1)

    
    celdaActualGTiVariante = hojaGTi.cell(row=rowActual, column= 3)  #ES LA COLUMNA 3 PORQUE CORRESPONDA AL AJ23 1 PRO 3 LHD - RHD
    print("buscando primer hilo de columna....") 
        
    while(True):
        if celdaActualGTiVariante.value == None:
            celdaActualGTiVariante = hojaGTi.cell(row=rowActual + 1, column=3)
            celdaPrimerHilo = celdaActualGTiVariante
            print("se encontro el primer hilo de la estacion, en la celda: " + str(celdaPrimerHilo))
            return celdaPrimerHilo            
        else:
            print("Hilo Actual: " + str(celdaActualGTiVariante.value) + " regresando un lugar atras")
            rowActual = rowActual - 1
            celdaActualGTiVariante = hojaGTi.cell(rowActual, column=3)
                
                
def validarHilos(celdaDelPrimerHilo, nameHojaAgrupcomb, nameHojaGTi):
    hojaAgrupcomb = fileAgrupcomb[nameHojaAgrupcomb]
    rowAgrup = 10
    celdaPrimerHiloAgrupcomb = hojaAgrupcomb.cell(row=rowAgrup, column= 1)
    while celdaPrimerHiloAgrupcomb.value != None:
        if celdaPrimerHiloAgrupcomb.value not in ["CIRCUITO POKA YOKE", "CIRCUITO AMG"]:
            print("Hilo Actual de Agrupcomb: "  + celdaPrimerHiloAgrupcomb.value)
            hiloActualAgrupcomb = str(celdaPrimerHiloAgrupcomb.value)
            
            if hiloActualAgrupcomb.startswith("EL-NO"):
                print("es un NO APLICA, SE SLTA AL SIGUIENTE HILO")
                rowAgrup += 1
                celdaPrimerHiloAgrupcomb = hojaAgrupcomb.cell(row=rowAgrup, column= 1)
            
            
            if hiloActualAgrupcomb.startswith("EL"):
                if hiloActualAgrupcomb.startswith("EL-NO"):
                    print("es un NO APLICA, SE SLTA AL SIGUIENTE HILO")
                else:
                    print("El Hilo es una tierra, buscando...")
                    validarTierras(hiloActualAgrupcomb,celdaDelPrimerHilo, nameHojaGTi)
            else:
                print("El hilo es circuito comun, buscando...")
                validarSerie(hiloActualAgrupcomb,celdaDelPrimerHilo, nameHojaGTi)
            
        rowAgrup += 1
        celdaPrimerHiloAgrupcomb = hojaAgrupcomb.cell(row=rowAgrup, column= 1)
    return        
    

sheetAgrupcomb = 'EAJ3-05'
sheetGTi = 'EST 1-8'   
    
    
    
    
    

print("Entrando a programa...")
print("\n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n BUSCANDO CELDA DE VARIANTE en GTi \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - -")            

celdaActualAgrVariante, celdaActualGTiVariante = buscarPrimeraCeldaDeVariante(sheetAgrupcomb,sheetGTi)  #busca la celda de la variante en el programa gti para posicionarse   


print("\n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n VALIDANDO CELDA DE VARIANTE \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - -")          

validarVarianteEnColumna(celdaActualGTiVariante,sheetAgrupcomb,'EST 1-8') #valida si todas las variantes de estacion estan en el programa gti

print("\n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n BUSCANDO FILA DE HILO \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - -")  

celdaDelPrimerHiloGTi = BuscarFilaDeHilo(celdaActualGTiVariante,sheetAgrupcomb,sheetGTi ) #busca la celda para pocisionarse en el programa gti

print("celda primer variante de estacion: " + str(celdaActualGTiVariante))
print("celda primer hilo de estacion: " + str(celdaDelPrimerHiloGTi))

print("\n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n VALIDANDO HILOS \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - -")  

validarHilos(celdaDelPrimerHiloGTi,sheetAgrupcomb,'EST 1-8')

print("hilos validados")

print("\n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n VALIDANDO CROSS MULTIPLE \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - - \n - - - - - - -")

aplicarCrossMultiple(celdaDelPrimerHiloGTi,celdaActualGTiVariante,sheetAgrupcomb,sheetGTi)



print("Finalizado")

fileGTi.save('C:/Users/sauceju/Desktop/Ruben Omar Aguero/open pyxl/prueba1/PRUEBA PROGRAMACION MOTOR.xlsx')
fileAgrupcomb.close()
fileGTi.close()